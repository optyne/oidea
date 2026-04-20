"""Generate pricing workbook (.xlsx) with 6 sheets and render each as PNG previews."""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent
OUT_XLSX = BASE / "報價系統.xlsx"
PREVIEW_DIR = BASE / "預覽圖"
PREVIEW_DIR.mkdir(exist_ok=True)

wb = Workbook()
wb.remove(wb.active)

# ---------- Shared styles ----------
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="微軟正黑體", size=16, bold=True, color="1F3864")
BODY_FONT = Font(name="微軟正黑體", size=10)
HILITE_FILL = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
TOTAL_FONT = Font(name="微軟正黑體", size=12, bold=True, color="C00000")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
BORDER = Border(*([Side(style="thin", color="BFBFBF")] * 4))
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

def style_body(ws, start_row, end_row, cols, zebra=True):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = LEFT
            cell.border = BORDER
            if zebra and (r - start_row) % 2 == 1:
                cell.fill = ALT_FILL

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ---------- Sheet 1: 使用說明 ----------
ws = wb.create_sheet("📘 使用說明")
ws["A1"] = "📘 使用說明"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:C1")
ws.row_dimensions[1].height = 32

ws["A3"] = "工作表導覽"
ws["A3"].font = Font(name="微軟正黑體", size=12, bold=True, color="1F3864")

rows = [
    ("順序", "工作表", "何時使用"),
    ("1", "📘 使用說明", "首次使用時閱讀（本頁）"),
    ("2", "💎 產品價目表", "新客戶報價前查方案與單價"),
    ("3", "📋 客戶主檔", "成交後登錄客戶基本資料"),
    ("4", "📦 服務項目", "記錄客戶已購買的服務明細"),
    ("5", "💵 收款紀錄 2026", "每次收到款項即登錄"),
    ("6", "🧮 報價單範例", "新客戶報價時複製本表改名使用"),
]
for i, row in enumerate(rows):
    for j, v in enumerate(row):
        ws.cell(row=4 + i, column=1 + j, value=v)
style_header(ws, 4, 3)
style_body(ws, 5, 10, 3)

ws["A12"] = "典型流程"
ws["A12"].font = Font(name="微軟正黑體", size=12, bold=True, color="1F3864")
flow = [
    "① 客戶詢問 → 打開「💎 產品價目表」確認方案",
    "② 複製「🧮 報價單範例」→ 重新命名為客戶名稱",
    "③ 填入方案、折扣、合計 → 寄出報價",
    "④ 客戶成交 → 「📋 客戶主檔」加一列",
    "⑤ 每個方案 → 「📦 服務項目」加一列",
    "⑥ 客戶付款 → 「💵 收款紀錄」加一列",
]
for i, line in enumerate(flow):
    c = ws.cell(row=13 + i, column=1, value=line)
    c.font = BODY_FONT
    ws.merge_cells(start_row=13 + i, start_column=1, end_row=13 + i, end_column=3)

set_widths(ws, [8, 24, 48])

# ---------- Sheet 2: 產品價目表 ----------
ws = wb.create_sheet("💎 產品價目表")
ws["A1"] = "💎 產品價目表"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:H1")
ws.row_dimensions[1].height = 32

headers = ["類別", "方案名稱", "產品代碼", "規格說明", "單位", "單價(NTD)", "主打", "備註"]
for i, h in enumerate(headers, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, len(headers))

products = [
    ("網站", "官網維護", "WEB-01", "月度例行維護、內容更新、小幅樣式調整、備份", "月", "[待填]", "", "不含大改版"),
    ("Email 主方案", "輕量版", "MAIL-LITE", "5GB 信箱 × 3 組", "月", "[待填]", "", "適合個人工作室"),
    ("Email 主方案", "標準版", "MAIL-STD", "20GB 信箱 × 10 組、垃圾郵件過濾", "月", "[待填]", "⭐", "最常成交方案"),
    ("Email 主方案", "專業版", "MAIL-PRO", "50GB 信箱 × 30 組、進階過濾、獨立 SMTP", "月", "[待填]", "", "中型公司適用"),
    ("容量加購", "小量", "CAP-S", "額外 10GB", "月", "[待填]", "", "加購於主方案"),
    ("容量加購", "中量", "CAP-M", "額外 50GB", "月", "[待填]", "", "加購於主方案"),
    ("容量加購", "海量", "CAP-L", "額外 200GB", "月", "[待填]", "", "加購於主方案"),
    ("別名服務", "單個別名", "ALIAS-01", "一個別名地址、永久", "月", "[待填]", "", ""),
    ("別名服務", "組合包", "ALIAS-03", "三個別名地址、永久", "月", "[待填]", "", "比單買省"),
    ("別名服務", "時效性", "ALIAS-T", "指定期間有效、到期自動停用", "次", "[待填]", "", "活動／短期用"),
]
for i, p in enumerate(products):
    for j, v in enumerate(p):
        ws.cell(row=4 + i, column=1 + j, value=v)
    if p[6] == "⭐":
        for c in range(1, len(headers) + 1):
            ws.cell(row=4 + i, column=c).fill = HILITE_FILL

style_body(ws, 4, 3 + len(products), len(headers), zebra=False)
set_widths(ws, [14, 14, 14, 38, 6, 12, 6, 20])

# ---------- Sheet 3: 客戶主檔 ----------
ws = wb.create_sheet("📋 客戶主檔")
ws["A1"] = "📋 客戶主檔"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:M1")
ws.row_dimensions[1].height = 32

headers = ["客戶編號", "客戶名稱", "統編", "聯絡人", "職稱", "電話", "行動電話", "Email", "寄件地址", "開始合作日", "客戶來源", "狀態", "備註"]
for i, h in enumerate(headers, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, len(headers))

customers = [
    ("C26001", "範例科技有限公司", "12345678", "王大明", "資訊主管", "02-1234-5678", "0912-345-678",
     "contact@example-tech.com.tw", "台北市中山區範例路1號", "2026-01-15", "朋友介紹", "活躍", "Email+官網"),
    ("C26002", "範例設計工作室", "", "李小美", "負責人", "", "0933-987-654",
     "hello@example-design.tw", "", "2026-03-01", "Google 搜尋", "活躍", "僅 Email 輕量版"),
]
for i, cust in enumerate(customers):
    for j, v in enumerate(cust):
        ws.cell(row=4 + i, column=1 + j, value=v)

style_body(ws, 4, 3 + len(customers), len(headers))
set_widths(ws, [10, 22, 10, 10, 10, 15, 14, 28, 22, 12, 12, 8, 18])

# ---------- Sheet 4: 服務項目 ----------
ws = wb.create_sheet("📦 服務項目")
ws["A1"] = "📦 服務項目"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:O1")
ws.row_dimensions[1].height = 32

headers = ["項目編號", "客戶編號", "客戶名稱", "產品代碼", "方案名稱", "數量",
           "單價(NTD)", "折扣%", "小計(NTD)", "計費週期", "起始日", "到期日", "自動續約", "狀態", "備註"]
for i, h in enumerate(headers, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, len(headers))

services = [
    ("S26001", "C26001", "範例科技有限公司", "WEB-01", "官網維護", 1, 8000, 0, None, "月", "2026-01-15", "2027-01-14", "是", "生效中", ""),
    ("S26002", "C26001", "範例科技有限公司", "MAIL-STD", "Email 標準版", 1, 3500, 0, None, "月", "2026-01-15", "2027-01-14", "是", "生效中", "主打方案"),
    ("S26003", "C26001", "範例科技有限公司", "CAP-M", "容量加購中量", 1, 1200, 10, None, "月", "2026-02-01", "2027-01-31", "是", "生效中", "老客戶折扣"),
    ("S26004", "C26002", "範例設計工作室", "MAIL-LITE", "Email 輕量版", 1, 1500, 0, None, "月", "2026-03-01", "2027-02-28", "是", "生效中", ""),
    ("S26005", "C26002", "範例設計工作室", "ALIAS-03", "別名組合包", 1, 500, 0, None, "月", "2026-03-01", "2027-02-28", "是", "生效中", ""),
]
for i, s in enumerate(services):
    for j, v in enumerate(s):
        cell = ws.cell(row=4 + i, column=1 + j)
        if j == 8:  # 小計 = 數量 × 單價 × (1 - 折扣%/100)
            r = 4 + i
            cell.value = f"=F{r}*G{r}*(1-H{r}/100)"
        else:
            cell.value = v
    if s[7] > 0:
        ws.cell(row=4 + i, column=8).fill = HILITE_FILL

style_body(ws, 4, 3 + len(services), len(headers))
set_widths(ws, [10, 10, 22, 12, 16, 6, 10, 8, 10, 8, 12, 12, 10, 10, 16])

# ---------- Sheet 5: 收款紀錄 2026 ----------
ws = wb.create_sheet("💵 收款紀錄 2026")
ws["A1"] = "💵 收款紀錄 2026"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:K1")
ws.row_dimensions[1].height = 32

headers = ["收款日期", "客戶編號", "客戶名稱", "項目編號", "內容摘要", "金額(NTD)",
           "付款方式", "發票號碼", "發票日期", "對帳狀態", "備註"]
for i, h in enumerate(headers, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, len(headers))

payments = [
    ("2026-01-20", "C26001", "範例科技有限公司", "S26001", "2026 Q1 官網維護", 24000, "銀行轉帳", "AB-12345678", "2026-01-20", "已對帳", ""),
    ("2026-01-20", "C26001", "範例科技有限公司", "S26002", "2026 Q1 Email 標準版", 10500, "銀行轉帳", "AB-12345679", "2026-01-20", "已對帳", ""),
    ("2026-03-05", "C26002", "範例設計工作室", "S26004", "2026 Q1 Email 輕量版", 4500, "LINE Pay", "AB-12345680", "2026-03-05", "已對帳", ""),
    ("2026-03-05", "C26002", "範例設計工作室", "S26005", "2026 Q1 別名組合包", 1500, "LINE Pay", "AB-12345681", "2026-03-05", "已對帳", ""),
]
for i, p in enumerate(payments):
    for j, v in enumerate(p):
        ws.cell(row=4 + i, column=1 + j, value=v)

style_body(ws, 4, 3 + len(payments), len(headers))

total_row = 4 + len(payments) + 1
ws.cell(row=total_row, column=5, value="年度合計").font = TOTAL_FONT
ws.cell(row=total_row, column=6, value=f"=SUM(F4:F{total_row-1})").font = TOTAL_FONT
for c in [5, 6]:
    ws.cell(row=total_row, column=c).fill = TOTAL_FILL
    ws.cell(row=total_row, column=c).alignment = CENTER

set_widths(ws, [12, 10, 22, 10, 22, 12, 12, 14, 12, 10, 14])

# ---------- Sheet 6: 報價單範例 ----------
ws = wb.create_sheet("🧮 報價單範例")
ws["A1"] = "🧮 報價單"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:E1")
ws.row_dimensions[1].height = 32

info = [
    ("", "", "", "報價編號", "Q26-0001"),
    ("", "", "", "報價日期", "2026-04-20"),
    ("", "", "", "有效期", "30 天"),
]
for i, row in enumerate(info):
    for j, v in enumerate(row):
        ws.cell(row=2 + i, column=1 + j, value=v)
    ws.cell(row=2 + i, column=4).font = Font(name="微軟正黑體", size=10, bold=True)

ws["A6"] = "客戶資訊"
ws["A6"].font = Font(name="微軟正黑體", size=12, bold=True, color="1F3864")
customer_fields = ["客戶名稱", "聯絡人", "電話", "Email"]
for i, f in enumerate(customer_fields):
    ws.cell(row=7 + i, column=1, value=f).font = Font(name="微軟正黑體", size=10, bold=True)
    ws.cell(row=7 + i, column=2, value="[填入]").font = BODY_FONT
    ws.merge_cells(start_row=7 + i, start_column=2, end_row=7 + i, end_column=5)

ws["A12"] = "報價明細"
ws["A12"].font = Font(name="微軟正黑體", size=12, bold=True, color="1F3864")
qheaders = ["產品代碼", "方案名稱", "數量", "單價(NTD)", "小計(NTD)"]
for i, h in enumerate(qheaders):
    ws.cell(row=13, column=1 + i, value=h)
style_header(ws, 13, 5)

sample_items = [
    ("MAIL-STD", "Email 標準版 × 3 個月", 3, 3500),
    ("WEB-01", "官網維護 × 3 個月", 3, 8000),
    ("ALIAS-03", "別名組合包 × 3 個月", 3, 500),
    ("", "", "", ""),
]
for i, it in enumerate(sample_items):
    r = 14 + i
    for j, v in enumerate(it):
        ws.cell(row=r, column=1 + j, value=v)
    ws.cell(row=r, column=5, value=f"=C{r}*D{r}" if it[2] else "")

style_body(ws, 14, 17, 5, zebra=False)

ws.cell(row=18, column=4, value="未稅小計").font = Font(name="微軟正黑體", size=10, bold=True)
ws.cell(row=18, column=5, value="=SUM(E14:E17)")
ws.cell(row=19, column=4, value="營業稅 5%").font = Font(name="微軟正黑體", size=10, bold=True)
ws.cell(row=19, column=5, value="=E18*0.05")
ws.cell(row=20, column=4, value="合計（含稅）").font = TOTAL_FONT
ws.cell(row=20, column=5, value="=E18+E19").font = TOTAL_FONT
for c in [4, 5]:
    ws.cell(row=20, column=c).fill = TOTAL_FILL
    ws.cell(row=20, column=c).alignment = CENTER

ws["A22"] = "備註"
ws["A22"].font = Font(name="微軟正黑體", size=12, bold=True, color="1F3864")
notes = [
    "1. 本報價含稅",
    "2. 付款方式：銀行轉帳／LINE Pay／現金",
    "3. 報價有效期內有效，逾期請重新確認",
    "4. 服務正式啟用日以收到款項後起算",
    "5. 續約請於到期前 30 日告知",
]
for i, n in enumerate(notes):
    c = ws.cell(row=23 + i, column=1, value=n)
    c.font = BODY_FONT
    ws.merge_cells(start_row=23 + i, start_column=1, end_row=23 + i, end_column=5)

set_widths(ws, [14, 28, 8, 14, 14])

wb.save(OUT_XLSX)
print(f"✅ 已生成 {OUT_XLSX}")
